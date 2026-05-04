import streamlit as st
import pandas as pd
from data_parser import load_data
import plotly.express as px
import io
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# Page configuration

# --- Custom CSS for Professional Look ---
st.markdown("""
<style>
    /* Main fonts */
    .stApp {
        font-family: 'Sarabun', 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
    }
    
    /* Hide Streamlit branding */
    #MainMenu {visibility: hidden;}
    footer {visibility: hidden;}
    header {visibility: hidden;}

    /* Style Metric Cards */
    div[data-testid="metric-container"] {
        border-radius: 10px;
        padding: 20px;
        box-shadow: 0 4px 6px rgba(0, 0, 0, 0.1);
        transition: transform 0.2s;
    }
    div[data-testid="metric-container"]:hover {
        transform: translateY(-5px);
        box-shadow: 0 6px 12px rgba(0, 0, 0, 0.15);
    }
    div[data-testid="metric-container"] > div {
        text-align: center;
    }
    div[data-testid="metric-container"] label {
        font-size: 1.1rem !important;
        font-weight: 600;
    }
    div[data-testid="metric-container"] div[data-testid="stMetricValue"] {
        font-size: 2.5rem !important;
        font-weight: bold;
    }

    /* Headers */
    h1, h2, h3 {
        font-weight: 700;
    }
    
    /* Table styling improvements */
    .stDataFrame {
        border-radius: 10px;
        overflow: hidden;
    }
</style>
""", unsafe_allow_html=True)

# --- Header Section ---
col_logo, col_title = st.columns([1, 10])
with col_logo:
    st.markdown('<div style="font-size: 50px; margin-top: 10px;">🏢</div>', unsafe_allow_html=True)
with col_title:
    st.title("ระบบตรวจสอบข้อมูลการลาบุคลากร")
    st.markdown("<p style='color: #666; font-size: 1.1rem;'>วิเคราะห์และสรุปผลข้อมูลการลาประจำปี (ป่วย, กิจ, พักผ่อน, สาย)</p>", unsafe_allow_html=True)

st.markdown("---")

# --- File Uploader ---
st.sidebar.markdown("### 📂 นำเข้าข้อมูล")
uploaded_file = st.sidebar.file_uploader("อัปโหลดไฟล์ Excel (.xlsx)", type=["xlsx"], help="รองรับไฟล์รูปแบบตารางวันลา")

@st.cache_data(show_spinner=False)
def get_data(file_source, _force_refresh=1):
    return load_data(file_source)

if uploaded_file is not None:
    with st.spinner('กำลังประมวลผลข้อมูล...'):
        df = get_data(uploaded_file, _force_refresh=2)
    
    if df is None or df.empty:
        st.error("❌ ไม่พบข้อมูลการลาในไฟล์ที่อัปโหลด หรือรูปแบบไฟล์ไม่ถูกต้อง")
    else:
        # --- Sidebar Filters ---
        st.sidebar.markdown("---")
        st.sidebar.markdown("### 🎯 ตัวกรองข้อมูล")
        
        budget_year = st.sidebar.multiselect(
            "📆 ปีงบประมาณ",
            options=sorted(df['BudgetYear'].unique()),
            default=df['BudgetYear'].unique()
        )
        
        months = st.sidebar.multiselect(
            "📅 เดือน",
            options=df['Month'].unique(),
            default=df['Month'].unique()
        )
        
        leave_types = st.sidebar.multiselect(
            "🏷️ ประเภทการลา",
            options=df['Type'].unique(),
            default=df['Type'].unique()
        )

        # Filter data
        filtered_df = df[
            df['BudgetYear'].isin(budget_year) & 
            df['Month'].isin(months) &
            df['Type'].isin(leave_types)
        ].copy()
        
        filtered_df['LeaveValue'] = filtered_df['IsHalf'].apply(lambda x: 0.5 if x else 1.0)

        # --- Metrics Summary ---
        st.markdown("### 📌 สรุปภาพรวม")
        col1, col2, col3, col4 = st.columns(4)
        
        def get_count(l_type):
            return filtered_df[filtered_df['Type'] == l_type]['LeaveValue'].sum()

        # Use emoji icons for metrics
        col1.metric("🤒 ลาป่วย", f"{get_count('ลาป่วย')} ครั้ง")
        col2.metric("🏃 ลากิจ", f"{get_count('ลากิจ')} ครั้ง")
        col3.metric("🏖️ ลาพักผ่อน", f"{get_count('ลาพักผ่อน')} ครั้ง")
        col4.metric("⏰ มาสาย", f"{get_count('สาย')} ครั้ง")
        
        st.markdown("<br>", unsafe_allow_html=True)

        # --- Main Content Tabs ---
        tab1, tab2 = st.tabs(["📋 รายการข้อมูลเชิงลึก", "📊 กราฟวิเคราะห์สถิติ"])

        with tab1:
            st.markdown("#### 📄 รายละเอียดการลาทั้งหมดตามตัวกรอง")
            display_df = filtered_df[['BudgetYear', 'Month', 'Day', 'Name', 'Type', 'IsHalf', 'Remarks']].copy()
            display_df.columns = ['ปีงบ', 'เดือน', 'วันที่', 'ชื่อ-สกุล', 'ประเภท', 'ครึ่งวัน', 'หมายเหตุ']
            
            # Format dataframe for better display
            st.dataframe(
                display_df, 
                use_container_width=True,
                height=400,
                hide_index=True
            )

        with tab2:
            st.markdown("#### 📈 เปรียบเทียบจำนวนครั้งการลาแยกตามบุคคล")
            summary_by_person = filtered_df.groupby(['Name', 'Type'])['LeaveValue'].sum().reset_index(name='จำนวนครั้ง')
            
            if not summary_by_person.empty:
                # Use professional color palette
                color_discrete_map = {
                    'ลาป่วย': '#ef553b', 
                    'ลากิจ': '#00cc96', 
                    'ลาพักผ่อน': '#636efa', 
                    'สาย': '#ab63fa'
                }
                
                fig = px.bar(
                    summary_by_person, 
                    x='Name', 
                    y='จำนวนครั้ง', 
                    color='Type',
                    barmode='group',
                    color_discrete_map=color_discrete_map,
                    template='plotly_white'
                )
                fig.update_layout(
                    xaxis_title="รายชื่อบุคลากร",
                    yaxis_title="จำนวนครั้ง",
                    legend_title="ประเภท",
                    font=dict(family="Sarabun, sans-serif"),
                    hovermode="x unified",
                    margin=dict(l=20, r=20, t=20, b=20)
                )
                st.plotly_chart(fig, use_container_width=True)
            else:
                st.info("ไม่พบข้อมูลสำหรับสร้างกราฟจากตัวกรองที่เลือก")

        # --- Batch Export Section ---
        def generate_batch_report(df_to_export):
            buffer = io.BytesIO()
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "ประวัติการลาทั้งหมด"
            
            font_bold = Font(bold=True, name='TH SarabunPSK', size=16)
            font_normal = Font(name='TH SarabunPSK', size=16)
            align_center = Alignment(horizontal='center', vertical='center')
            align_left = Alignment(horizontal='left', vertical='center')
            border = Border(
                left=Side(border_style='thin', color='000000'),
                right=Side(border_style='thin', color='000000'),
                top=Side(border_style='thin', color='000000'),
                bottom=Side(border_style='thin', color='000000')
            )
            fill_yellow = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
            fill_purple = PatternFill(start_color='C0A0C0', end_color='C0A0C0', fill_type='solid')
            
            row_num = 1
            months_order = ['ตุลาคม', 'พฤศจิกายน', 'ธันวาคม', 'มกราคม', 'กุมภาพันธ์', 'มีนาคม', 'เมษายน', 'พฤษภาคม', 'มิถุนายน', 'กรกฎาคม', 'สิงหาคม', 'กันยายน']
            
            for person in sorted(df_to_export['Name'].unique()):
                p_data = df_to_export[df_to_export['Name'] == person].sort_values(['BudgetYear', 'Month', 'Day'])
                if p_data.empty: continue
                
                budget_years = p_data['BudgetYear'].unique()
                target_year = budget_years[0] if len(budget_years) > 0 else '2568'
                prev_year = int(target_year) - 1
                
                ws.cell(row=row_num, column=1, value=f"สรุปวันลาปีงบประมาณ (1 ตุลาคม {prev_year} - 30 กันยายน {target_year})").font = font_bold
                ws.cell(row=row_num, column=1).alignment = align_center
                ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=6)
                row_num += 1
                
                ws.cell(row=row_num, column=1, value=person).font = font_bold
                ws.cell(row=row_num, column=1).alignment = align_center
                ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=6)
                row_num += 1
                
                ws.cell(row=row_num, column=1, value="มีสิทธิลาป่วยโดยไม่ถูกหักค่าจ้าง 15 วัน").font = font_bold
                ws.cell(row=row_num, column=1).alignment = align_center
                ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=6)
                row_num += 1
                
                headers = ['วัน เดือน ปี', 'ป่วย', 'กิจ', 'พักผ่อน', 'สาย', 'หมายเหตุ']
                for col_num, header in enumerate(headers, 1):
                    cell = ws.cell(row=row_num, column=col_num, value=header)
                    cell.font = font_bold
                    cell.alignment = align_center
                    cell.border = border
                row_num += 1
                
                sums = {'ลาป่วย': 0.0, 'ลากิจ': 0.0, 'ลาพักผ่อน': 0.0, 'สาย': 0.0}
                
                for _, row in p_data.iterrows():
                    month = row['Month']
                    month_idx = months_order.index(month) if month in months_order else 0
                    year_str = str(prev_year) if month_idx < 3 else str(target_year)
                    date_str = f"วันที่ {row['Day']} {month} {year_str}"
                    
                    ws.cell(row=row_num, column=1, value=date_str).alignment = align_center
                    
                    type_str = row['Type']
                    is_half = row['IsHalf']
                    val_str = 'ครึ่ง' if is_half else 1
                    val_num = 0.5 if is_half else 1.0
                    
                    col_map = {'ลาป่วย': 2, 'ลากิจ': 3, 'ลาพักผ่อน': 4, 'สาย': 5}
                    target_col = col_map.get(type_str, 2)
                    if type_str in sums: sums[type_str] += val_num
                    
                    ws.cell(row=row_num, column=target_col, value=val_str).alignment = align_center
                    for c in range(2, 6):
                        if c != target_col: ws.cell(row=row_num, column=c, value='')
                    ws.cell(row=row_num, column=6, value=row['Remarks']).alignment = align_left
                    
                    for c in range(1, 7):
                        ws.cell(row=row_num, column=c).border = border
                        ws.cell(row=row_num, column=c).font = font_normal
                    row_num += 1
                
                ws.cell(row=row_num, column=1, value='รวมทั้งปีงบประมาณ').alignment = align_center
                ws.cell(row=row_num, column=1).font = font_bold
                ws.cell(row=row_num, column=1).border = border
                ws.cell(row=row_num, column=1).fill = fill_yellow
                
                def format_sum(s):
                    if s == 0: return ""
                    int_part = int(s)
                    if s == int_part: return str(int_part)
                    if int_part == 0: return "ครึ่ง"
                    return f"{int_part}ครึ่ง"
                
                for i, key in enumerate(['ลาป่วย', 'ลากิจ', 'ลาพักผ่อน', 'สาย']):
                    cell = ws.cell(row=row_num, column=i+2, value=format_sum(sums[key]))
                    cell.alignment = align_center
                    cell.font = font_bold
                    cell.border = border
                    cell.fill = fill_purple
                
                ws.cell(row=row_num, column=6, value='').border = border
                row_num += 1

            ws.column_dimensions['A'].width = 25
            ws.column_dimensions['B'].width = 8
            ws.column_dimensions['C'].width = 8
            ws.column_dimensions['D'].width = 12
            ws.column_dimensions['E'].width = 8
            ws.column_dimensions['F'].width = 30
            
            wb.save(buffer)
            return buffer.getvalue()

        st.markdown("<br>", unsafe_allow_html=True)
        col_batch, col_graph, _ = st.columns([1.5, 1.5, 1])
        with col_batch:
            st.download_button(
                label="📥 ดาวน์โหลดรายงานรายบุคคลแบบทั้งหมด (Excel)",
                data=generate_batch_report(filtered_df),
                file_name="รายงานการลา_ทั้งหมด.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                type="primary"
            )
            
        with col_graph:
            # Generate summary data for graph export
            graph_data = filtered_df.groupby(['Name', 'Type'])['LeaveValue'].sum().reset_index(name='จำนวนวัน/ครั้ง')
            if not graph_data.empty:
                pivot_graph = graph_data.pivot_table(index='Name', columns='Type', values='จำนวนวัน/ครั้ง', fill_value=0).reset_index()
                buffer_graph = io.BytesIO()
                with pd.ExcelWriter(buffer_graph, engine='openpyxl') as writer:
                    pivot_graph.to_excel(writer, index=False, sheet_name='ข้อมูลกราฟ')
                
                st.download_button(
                    label="📊 ดาวน์โหลดข้อมูลกราฟ (Excel)",
                    data=buffer_graph.getvalue(),
                    file_name="สรุปข้อมูลกราฟ.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )

        # --- Individual Search Section ---
        st.markdown("---")
        st.markdown("### 🔍 เจาะลึกรายบุคคล")
        
        col_search, col_space = st.columns([1, 2])
        with col_search:
            selected_person = st.selectbox(
                "👤 เลือกระบุรายชื่อเพื่อดูประวัติทั้งหมด", 
                options=["-- เลือกรายชื่อ --"] + sorted(df['Name'].unique())
            )
        
        if selected_person != "-- เลือกรายชื่อ --":
            # Use filtered_df instead of df so sidebar filters apply here too
            person_data = filtered_df[filtered_df['Name'] == selected_person].sort_values(['BudgetYear', 'Month', 'Day'])
            
            st.markdown(f"##### ประวัติของ: <span style='color:#1f77b4;'>{selected_person}</span>", unsafe_allow_html=True)
            
            if person_data.empty:
                st.warning("ไม่พบประวัติการลา")
            else:
                # --- Calculate Detailed Summary ---
                st.markdown("##### 📊 สรุปยอดการลา (แยกตามเดือน)")
                
                # Group by BudgetYear, Month, and Type
                monthly_summary = person_data.groupby(['BudgetYear', 'Month', 'Type'])['LeaveValue'].sum().reset_index()
                
                # Create a pivot table for better display: Rows = Month, Columns = Leave Type
                pivot_summary = monthly_summary.pivot_table(
                    index=['BudgetYear', 'Month'], 
                    columns='Type', 
                    values='LeaveValue', 
                    fill_value=0
                ).reset_index()
                
                # Display the summary table
                st.dataframe(
                    pivot_summary.style.format(precision=1),
                    use_container_width=True,
                    hide_index=True
                )
                
                # Total summary for the person
                st.markdown("##### 📌 สรุปยอดรวมทั้งหมด")
                total_summary = person_data.groupby('Type')['LeaveValue'].sum().reset_index()
                cols = st.columns(len(total_summary) if not total_summary.empty else 1)
                for i, row in total_summary.iterrows():
                    cols[i].metric(row['Type'], f"{row['LeaveValue']} วัน/ครั้ง")
                
                st.markdown("<br>", unsafe_allow_html=True)
                
                budget_years = person_data['BudgetYear'].unique()
                target_year = budget_years[0] if len(budget_years) > 0 else '2568'
                prev_year = int(target_year) - 1
                
                # --- Detailed History Table ---
                st.markdown("##### 📄 ประวัติการลาโดยละเอียด")
                
                ui_display_data = []
                months_order = ['ตุลาคม', 'พฤศจิกายน', 'ธันวาคม', 'มกราคม', 'กุมภาพันธ์', 'มีนาคม', 'เมษายน', 'พฤษภาคม', 'มิถุนายน', 'กรกฎาคม', 'สิงหาคม', 'กันยายน']
                
                for _, row in person_data.iterrows():
                    month = row['Month']
                    month_idx = months_order.index(month) if month in months_order else 0
                    year_str = str(prev_year) if month_idx < 3 else str(target_year)
                    date_str = f"วันที่ {row['Day']} {month} {year_str}"
                    
                    type_str = row['Type']
                    is_half = row['IsHalf']
                    val_str = 'ครึ่ง' if is_half else '1'
                    
                    ui_row = {
                        'วัน เดือน ปี': date_str,
                        'ป่วย': val_str if type_str == 'ลาป่วย' else '',
                        'กิจ': val_str if type_str == 'ลากิจ' else '',
                        'พักผ่อน': val_str if type_str == 'ลาพักผ่อน' else '',
                        'สาย': val_str if type_str == 'สาย' else '',
                        'หมายเหตุ': row['Remarks']
                    }
                    ui_display_data.append(ui_row)
                
                ui_df = pd.DataFrame(ui_display_data)
                
                def style_center(val):
                    return 'text-align: center;'
                    
                st.dataframe(
                    ui_df.style.map(style_center, subset=['ป่วย', 'กิจ', 'พักผ่อน', 'สาย']) if hasattr(ui_df.style, "map") else ui_df.style.applymap(style_center, subset=['ป่วย', 'กิจ', 'พักผ่อน', 'สาย']),
                    use_container_width=True, 
                    hide_index=True
                )
                
                # --- Export to Excel ---
                st.markdown("<br>", unsafe_allow_html=True)
                
                buffer = io.BytesIO()
                
                # We will build the Excel file from scratch using openpyxl instead of pandas
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "ประวัติการลา"
                
                # Styles
                font_bold = Font(bold=True, name='TH SarabunPSK', size=16)
                font_normal = Font(name='TH SarabunPSK', size=16)
                align_center = Alignment(horizontal='center', vertical='center')
                align_left = Alignment(horizontal='left', vertical='center')
                border = Border(
                    left=Side(border_style='thin', color='000000'),
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='000000'),
                    bottom=Side(border_style='thin', color='000000')
                )
                
                # Row 1: สรุปวันลาปีงบประมาณ
                ws['A1'] = f"สรุปวันลาปีงบประมาณ (1 ตุลาคม {prev_year} - 30 กันยายน {target_year})"
                ws['A1'].font = font_bold
                ws['A1'].alignment = align_center
                ws.merge_cells('A1:F1')
                
                # Row 2: ชื่อ-สกุล
                ws['A2'] = selected_person
                ws['A2'].font = font_bold
                ws['A2'].alignment = align_center
                ws.merge_cells('A2:F2')
                
                # Row 3: มีสิทธิลาป่วยโดยไม่ถูกหักค่าจ้าง 15 วัน
                ws['A3'] = "มีสิทธิลาป่วยโดยไม่ถูกหักค่าจ้าง 15 วัน"
                ws['A3'].font = font_bold
                ws['A3'].alignment = align_center
                ws.merge_cells('A3:F3')
                
                # Row 4: Headers
                headers = ['วัน เดือน ปี', 'ป่วย', 'กิจ', 'พักผ่อน', 'สาย', 'หมายเหตุ']
                for col_num, header in enumerate(headers, 1):
                    cell = ws.cell(row=4, column=col_num)
                    cell.value = header
                    cell.font = font_bold
                    cell.alignment = align_center
                    cell.border = border
                    
                # Column widths
                ws.column_dimensions['A'].width = 25
                ws.column_dimensions['B'].width = 8
                ws.column_dimensions['C'].width = 8
                ws.column_dimensions['D'].width = 12
                ws.column_dimensions['E'].width = 8
                ws.column_dimensions['F'].width = 30
                
                # Data Rows
                row_num = 5
                sums = {'ลาป่วย': 0.0, 'ลากิจ': 0.0, 'ลาพักผ่อน': 0.0, 'สาย': 0.0}
                months_order = ['ตุลาคม', 'พฤศจิกายน', 'ธันวาคม', 'มกราคม', 'กุมภาพันธ์', 'มีนาคม', 'เมษายน', 'พฤษภาคม', 'มิถุนายน', 'กรกฎาคม', 'สิงหาคม', 'กันยายน']
                
                for _, row in person_data.iterrows():
                    month = row['Month']
                    month_idx = months_order.index(month) if month in months_order else 0
                    year_str = str(prev_year) if month_idx < 3 else str(target_year)
                    
                    date_str = f"วันที่ {row['Day']} {month} {year_str}"
                    ws.cell(row=row_num, column=1, value=date_str).alignment = align_center
                    
                    type_str = row['Type']
                    is_half = row['IsHalf']
                    val_str = 'ครึ่ง' if is_half else 1
                    val_num = 0.5 if is_half else 1.0
                    
                    col_map = {'ลาป่วย': 2, 'ลากิจ': 3, 'ลาพักผ่อน': 4, 'สาย': 5}
                    target_col = col_map.get(type_str, 2)
                    
                    if type_str in sums:
                        sums[type_str] += val_num
                        
                    ws.cell(row=row_num, column=target_col, value=val_str).alignment = align_center
                    
                    # Fill other columns with empty
                    for c in range(2, 6):
                        if c != target_col:
                            ws.cell(row=row_num, column=c, value='')
                            
                    ws.cell(row=row_num, column=6, value=row['Remarks']).alignment = align_left
                    
                    # Apply borders
                    for c in range(1, 7):
                        ws.cell(row=row_num, column=c).border = border
                        ws.cell(row=row_num, column=c).font = font_normal
                        
                    row_num += 1
                    
                # Summary Row
                ws.cell(row=row_num, column=1, value='รวมทั้งปีงบประมาณ').alignment = align_center
                ws.cell(row=row_num, column=1).font = font_bold
                ws.cell(row=row_num, column=1).border = border
                
                def format_sum(s):
                    if s == 0: return ""
                    int_part = int(s)
                    if s == int_part: return str(int_part)
                    if int_part == 0: return "ครึ่ง"
                    return f"{int_part}ครึ่ง"
                    
                for i, key in enumerate(['ลาป่วย', 'ลากิจ', 'ลาพักผ่อน', 'สาย']):
                    ws.cell(row=row_num, column=i+2, value=format_sum(sums[key])).alignment = align_center
                    ws.cell(row=row_num, column=i+2).font = font_bold
                    ws.cell(row=row_num, column=i+2).border = border
                    
                ws.cell(row=row_num, column=6, value='').border = border
                
                # Fill colors for summary (yellow for total label, purple for sum cols 2,3,4)
                fill_yellow = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                fill_purple = PatternFill(start_color='C0A0C0', end_color='C0A0C0', fill_type='solid')
                
                ws.cell(row=row_num, column=1).fill = fill_yellow
                ws.cell(row=row_num, column=2).fill = fill_purple
                ws.cell(row=row_num, column=3).fill = fill_purple
                ws.cell(row=row_num, column=4).fill = fill_purple
                ws.cell(row=row_num, column=5).fill = fill_purple
                
                wb.save(buffer)
                
                col_btn1, col_btn2 = st.columns([1, 3])
                with col_btn1:
                    st.download_button(
                        label="📥 ดาวน์โหลดรายงาน (Excel)",
                        data=buffer.getvalue(),
                        file_name=f"รายงานการลา_{selected_person.replace(' ', '_')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True
                    )

else:
    # Empty State UI
    st.markdown("<br><br>", unsafe_allow_html=True)
    col_img1, col_img2, col_img3 = st.columns([1, 2, 1])
    with col_img2:
        st.info("👋 ยินดีต้อนรับ! กรุณาอัปโหลดไฟล์ Excel ทางเมนูด้านซ้ายเพื่อเริ่มต้นใช้งานระบบ")
        st.markdown(
            """
            <div style='text-align: center; opacity: 0.5; margin-top: 20px;'>
                <img src="https://img.icons8.com/clouds/200/000000/upload.png" width="150" />
                <p>รองรับไฟล์ .xlsx</p>
            </div>
            """, 
            unsafe_allow_html=True
        )
